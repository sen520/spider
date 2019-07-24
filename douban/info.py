import requests
from lxml import etree
from openpyxl import Workbook
import json
import re
from tools.tools import data_to_json

def get_page(url, name):
    id = re.search('\/(\d+)\/', url)[1]
    res = requests.get(url)
    html = etree.HTML(res.text)
    t = html.xpath('//h1//text()')
    img = html.xpath('string(//a[@class="nbgnbg"]/img/@src)')
    alias = ''.join(list(map(lambda x: x.strip(), t)))
    abstract = html.xpath('string(//div[@id="info"])')
    intros = html.xpath('//div[@class="indent"]//text()')
    intro = '\n'.join(list(map(lambda x: x.strip(), intros)))
    actors = html.xpath('//li[@class="celebrity"]')
    actor_list = []
    for actor in actors:
        actor_website = actor.xpath('./a/@href')[0]
        actor_name, actor_title = actor.xpath('./div/span//text()')
        actor_list.append({'name': actor_name, 'title': actor_title, 'website': actor_website})
    star_div = html.xpath('//div[@id="interest_sectl"]')[0]
    star_score = star_div.xpath('string(.//strong/text())')
    star_num = star_div.xpath('string(.//div[@class="rating_sum"]//span/text())')

    hot_comments = html.xpath('//div[@id="hot-comments"]/div[@class="comment-item"]/div[@class="comment"]')
    comment_list = []
    for comment in hot_comments:
        commenter = comment.xpath('string(.//span[@class="comment-info"]/a/text())')
        comment_start = comment.xpath('string(.//span[@class="comment-info"]/span[2]/@title)')
        comment_text = comment.xpath('string(.//p/span/text())')
        comment_list.append({'commenter': commenter, 'comment_start': comment_start, 'comment_text': comment_text})

    auth_url = 'https://m.douban.com/rexxar/api/v2/movie/{}/verify_users?start=0&count=2&ck='.format(id)
    headers = {
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Origin': 'https://movie.douban.com',
        'Referer': 'https://movie.douban.com/subject/27191492/?from=subject-page',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36'
    }
    auth_res = requests.get(auth_url, headers=headers)
    auth_dict = json.loads(auth_res.text)
    try:
        auth_name = auth_dict['verify_users'][0]['user']['name']
        auth_intro =auth_dict['verify_users'][0]['articles'][0]['abstract']
        auth_abstract =auth_dict['verify_users'][0]['user']['abstract']
    except:
        auth_name = ''
        auth_intro = ''
        auth_abstract = ''
    movie_dict = {
        'name': name,
        'img': img,
        'alias': alias,
        'abstract': abstract,
        'intro': intro,
        'actors': actor_list,
        'stars': {
            'score': star_score,
            'num': star_num,
        },
        'auth': {
            'name': auth_name,
            'abstract': auth_abstract,
            'intro': auth_intro,
        },
        'comments': comment_list,
    }
    return movie_dict

def spider(url):
    res = requests.get(url)
    html = etree.HTML(res.text)
    tag_as = html.xpath('//tr[@class="item"]/td[2]//a')
    movie_list = []
    for tag_a in tag_as:
        name_list = tag_a.xpath('.//text()')
        name = ''.join(list(map(lambda x: x.replace(' ', '').replace('\n', '').strip(), name_list)))
        href = tag_a.xpath('string(./@href)')
        movie_dict = get_page(href, name)
        movie_list.append(movie_dict)
    data_to_json(movie_list, 'movie')


def dictTocsv(data):
    book = Workbook()
    sheet = book.active
    columns = data.keys()
    for index, c in enumerate(columns):
        sheet.cell(1, index + 1).value = c
    row = 2

    for j, c in enumerate(columns):
        sheet.cell(row, j + 1).value = data[c]
        row = row + 1
    book.save(u'xxx.xlsx')


if __name__ == '__main__':
    url = 'https://movie.douban.com/chart'
    spider(url)
